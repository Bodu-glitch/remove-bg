"""
Pipeline nhận diện tên sản phẩm từ ảnh + match với Excel:
1. qwen2.5vl (Ollama) phân tích ảnh → mô tả sản phẩm
2. rapidfuzz pre-filter top-10 ứng viên từ tvp.xlsx
3. qwen3:4b chọn tên khớp nhất trong top-10
4. Pillow tạo ảnh nền trắng + tên sản phẩm phía trên

Yêu cầu: ollama serve  (chạy trong terminal khác)
         ollama pull qwen2.5vl
         ollama pull qwen3:4b
         pip install openpyxl rapidfuzz

Chạy thử 5 ảnh: python label_product.py --test 5
Bỏ qua đã xử lý: python label_product.py --skip-done
Chạy tất cả:    python label_product.py
"""

import argparse
import base64
import csv
import io
import re
import sys
from pathlib import Path

import requests
from PIL import Image, ImageDraw, ImageFont
from rapidfuzz import fuzz, process, utils
import openpyxl

# ── Cấu hình ──────────────────────────────────────────────────────────────────

OLLAMA_URL   = "http://localhost:11434"
VISION_MODEL = "qwen2.5vl"   # nhận diện ảnh
TEXT_MODEL   = "qwen2.5vl"   # match tên (dùng chung với vision, không có thinking mode)

EXCEL_FILE   = Path("tvp.xlsx")
BRAND_FILE   = Path("brands.xlsx")     # 1 cột tên brand; bỏ qua nếu không có
INPUT_DIR    = Path("img_no_bg_dev")   # dev; đổi thành img_no_bg khi production
OUTPUT_DIR   = Path("img_labeled")
OUTPUT_CSV   = Path("product_mapping.csv")
UNMATCHED_CSV = Path("unmatched.csv")

OUTPUT_SIZE         = (1024, 1024)
PRODUCT_HEIGHT_RATIO = 0.72   # sản phẩm chiếm 72% chiều cao canvas
TEXT_AREA_HEIGHT    = 110     # px dành cho chữ phía trên sản phẩm

FONT_PATHS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
]
FONT_SIZE  = 32
TEXT_COLOR = (30, 30, 30)
BG_COLOR   = (255, 255, 255)

SUPPORTED_EXT = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_font(size: int = FONT_SIZE) -> ImageFont.FreeTypeFont:
    for path in FONT_PATHS:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def check_ollama(models: list[str]):
    try:
        r = requests.get(f"{OLLAMA_URL}/api/tags", timeout=5)
        available = [m["name"] for m in r.json().get("models", [])]
    except requests.exceptions.ConnectionError:
        print("LỖI: Ollama chưa chạy. Khởi động bằng: ollama serve")
        sys.exit(1)

    for model in models:
        base = model.split(":")[0]
        if not any(base in m for m in available):
            print(f"LỖI: Model '{model}' chưa được pull.")
            print(f"Chạy: ollama pull {model}")
            sys.exit(1)
    print(f"Ollama OK — {', '.join(models)} sẵn sàng.")


def load_product_names() -> list[str]:
    """Đọc tên sản phẩm từ cột B (index 1), bắt đầu dòng 5, tất cả sheets."""
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    names = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=5, values_only=True):
            val = row[1] if len(row) > 1 else None
            if val and str(val).strip():
                names.append(str(val).strip())
    wb.close()
    print(f"Đã tải {len(names)} tên sản phẩm từ {EXCEL_FILE} ({len(wb.sheetnames)} sheets)")
    return names


def load_brands() -> list[str]:
    if not BRAND_FILE.exists():
        return []
    wb = openpyxl.load_workbook(BRAND_FILE, read_only=True)
    brands = [row[0].value for row in wb.active.iter_rows() if row and row[0].value and str(row[0].value).strip()]
    wb.close()
    print(f"Đã tải {len(brands)} brand từ {BRAND_FILE}")
    return brands


# ── Step 1: qwen2.5vl phân tích ảnh ──────────────────────────────────────────

def analyze_image(image_path: Path) -> str:
    img = Image.open(image_path).convert("RGBA")

    # Paste lên nền trắng để model nhận diện rõ hơn
    preview = Image.new("RGB", (512, 512), (255, 255, 255))
    thumb = img.copy()
    thumb.thumbnail((512, 512), Image.LANCZOS)
    offset = ((512 - thumb.width) // 2, (512 - thumb.height) // 2)
    preview.paste(thumb, offset, mask=thumb.split()[3])

    buf = io.BytesIO()
    preview.save(buf, format="JPEG", quality=85)
    image_b64 = base64.b64encode(buf.getvalue()).decode()

    response = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={
            "model": VISION_MODEL,
            "prompt": (
                "Đây là ảnh sản phẩm trên nền trắng. "
                "Mô tả ngắn gọn đây là sản phẩm gì bằng tiếng Việt, "
                "tối đa 15 từ, không giải thích thêm."
            ),
            "images": [image_b64],
            "stream": False,
            "keep_alive": "5m",
            "options": {"temperature": 0.1, "num_predict": 40},
        },
        timeout=120,
    )
    return response.json().get("response", "").strip()


def detect_brand(image_path: Path, brands: list[str]) -> str | None:
    """Nhận diện brand có in trên sản phẩm/bao bì. Trả về tên brand hoặc None."""
    if not brands:
        return None

    img = Image.open(image_path).convert("RGBA")
    preview = Image.new("RGB", (512, 512), (255, 255, 255))
    thumb = img.copy()
    thumb.thumbnail((512, 512), Image.LANCZOS)
    offset = ((512 - thumb.width) // 2, (512 - thumb.height) // 2)
    preview.paste(thumb, offset, mask=thumb.split()[3])

    buf = io.BytesIO()
    preview.save(buf, format="JPEG", quality=85)
    image_b64 = base64.b64encode(buf.getvalue()).decode()

    brands_block = "\n".join(f"- {b}" for b in brands)
    prompt = (
        f"Nhìn vào ảnh sản phẩm, tìm tên thương hiệu (brand) in trên sản phẩm hoặc bao bì.\n"
        f"Danh sách brand:\n{brands_block}\n\n"
        "Nếu thấy brand nào trong danh sách, trả lời đúng tên brand đó.\n"
        "Nếu không thấy hoặc không rõ, trả lời: không rõ"
    )
    response = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={
            "model": VISION_MODEL,
            "prompt": prompt,
            "images": [image_b64],
            "stream": False,
            "keep_alive": "5m",
            "options": {"temperature": 0.0, "num_predict": 20},
        },
        timeout=120,
    )
    result = response.json().get("response", "").strip()
    print(f"    [brand-raw] '{result}'")
    result_lower = result.lower()
    for brand in brands:
        if brand.lower() in result_lower:
            print(f"    [brand-match] '{brand}'")
            return brand
    print(f"    [brand-nomatch] không tìm thấy brand nào trong response")
    return None


# ── Step 2: rapidfuzz + qwen3 match tên ──────────────────────────────────────

def match_product_name(description: str, product_names: list[str], brand: str | None = None) -> str | None:
    """Trả về tên khớp nhất hoặc None nếu không tìm được."""

    # Tầng 1: rapidfuzz pre-filter (lowercase + strip để fix case mismatch)
    candidates = process.extract(
        description, product_names,
        scorer=fuzz.WRatio,
        processor=utils.default_process,
        limit=20,
    )
    print(f"    [fuzzy-query] '{description}'")
    print(f"    [fuzzy-top5] " + " | ".join(f"'{n}'({s:.0f})" for n, s, _ in candidates[:5]))

    # Tầng 2: qwen3 chọn trong top-10
    names_block = "\n".join(f"{i+1}. {name}" for i, (name, _, _) in enumerate(candidates))
    brand_line = f"Brand: {brand}\n" if brand else ""
    prompt = (
        f"{brand_line}"
        f"Mô tả sản phẩm: {description}\n\n"
        f"Các tên sản phẩm có thể khớp:\n{names_block}\n\n"
        "Chọn số thứ tự của tên sản phẩm khớp nhất với mô tả. "
        "Nếu không có cái nào phù hợp, trả lời 0. "
        "Chỉ trả lời đúng 1 số nguyên, không giải thích."
    )

    response = requests.post(
        f"{OLLAMA_URL}/api/generate",
        json={
            "model": TEXT_MODEL,
            "prompt": prompt,
            "stream": False,
            "keep_alive": "5m",
            "options": {"temperature": 0.0, "num_predict": 50},
        },
        timeout=60,
    )
    data = response.json()
    raw = data.get("response", "").strip()
    # qwen3.5 đặt câu trả lời trong "thinking" khi response rỗng
    if not raw:
        raw = data.get("thinking", "").strip()
    print(f"    [qwen3-prompt]\n{prompt}\n---")
    print(f"    [qwen3-raw] '{raw[:200]}'")

    num = _parse_qwen3_number(raw, len(candidates))
    print(f"    [qwen3-parsed] num={num}")
    if num is None:
        return None

    return candidates[num - 1][0]


def _parse_qwen3_number(raw: str, max_idx: int) -> int | None:
    """
    Parse số từ response qwen3.
    - Nếu có <think> block: bỏ qua, lấy phần sau
    - Nếu dùng thinking field: tìm số cuối nằm trong range hợp lệ (1..max_idx)
    """
    clean = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()
    if not clean:
        clean = raw

    # Lấy tất cả số trong range hợp lệ, lấy số CUỐI (kết luận của thinking)
    numbers = [int(n) for n in re.findall(r"\b\d+\b", clean) if 0 <= int(n) <= max_idx]
    if not numbers:
        return None
    num = numbers[-1]
    if num < 1:  # 0 = model nói không khớp
        return None
    return num


# ── Step 3: Pillow tạo ảnh có chữ ────────────────────────────────────────────

def render_labeled_image(image_path: Path, product_name: str, output_path: Path):
    img = Image.open(image_path).convert("RGBA")
    canvas = Image.new("RGB", OUTPUT_SIZE, BG_COLOR)

    # Tính kích thước sản phẩm (chiếm PRODUCT_HEIGHT_RATIO của chiều cao)
    max_h = int(OUTPUT_SIZE[1] * PRODUCT_HEIGHT_RATIO)
    max_w = OUTPUT_SIZE[0] - 40  # margin 20px mỗi bên

    thumb = img.copy()
    thumb.thumbnail((max_w, max_h), Image.LANCZOS)

    # Căn giữa ngang, đặt phía dưới (cách bottom 40px)
    x = (OUTPUT_SIZE[0] - thumb.width) // 2
    y = OUTPUT_SIZE[1] - thumb.height - 40

    canvas.paste(thumb, (x, y), mask=thumb.split()[3])

    # Vẽ tên sản phẩm phía trên
    draw = ImageDraw.Draw(canvas)
    font = load_font(FONT_SIZE)

    # Auto-wrap nếu tên quá dài
    max_text_w = OUTPUT_SIZE[0] - 60
    lines = _wrap_text(product_name, font, max_text_w)

    line_h = FONT_SIZE + 6
    total_text_h = len(lines) * line_h
    text_y = (TEXT_AREA_HEIGHT - total_text_h) // 2 + (y - TEXT_AREA_HEIGHT) // 2 + 20

    for i, line in enumerate(lines):
        bbox = draw.textbbox((0, 0), line, font=font)
        text_w = bbox[2] - bbox[0]
        text_x = (OUTPUT_SIZE[0] - text_w) // 2
        draw.text((text_x, text_y + i * line_h), line, fill=TEXT_COLOR, font=font)

    canvas.save(output_path, format="PNG", optimize=True)


def _wrap_text(text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    """Chia dòng nếu text quá dài."""
    words = text.split()
    lines = []
    current = ""

    # Dùng ImageDraw tạm để đo text
    tmp = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    for word in words:
        test = (current + " " + word).strip()
        bbox = tmp.textbbox((0, 0), test, font=font)
        if bbox[2] - bbox[0] <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or [text]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Label sản phẩm từ ảnh + Excel")
    parser.add_argument("--test", type=int, metavar="N", help="Chỉ xử lý N ảnh đầu tiên")
    parser.add_argument("--skip-done", action="store_true", help="Bỏ qua ảnh đã có trong output")
    args = parser.parse_args()

    # Kiểm tra prerequisites
    if not EXCEL_FILE.exists():
        print(f"LỖI: Không tìm thấy {EXCEL_FILE}")
        sys.exit(1)
    if not INPUT_DIR.exists():
        print(f"LỖI: Thư mục {INPUT_DIR} không tồn tại")
        sys.exit(1)

    check_ollama([VISION_MODEL, TEXT_MODEL])
    product_names = load_product_names()
    brands = load_brands()

    OUTPUT_DIR.mkdir(exist_ok=True)

    # Lấy danh sách ảnh
    images = sorted([
        p for p in INPUT_DIR.iterdir()
        if p.suffix.lower() in SUPPORTED_EXT
    ])
    if args.test:
        images = images[: args.test]

    print(f"Xử lý {len(images)} ảnh từ {INPUT_DIR}/\n")

    matched_rows = []
    unmatched_rows = []

    for i, img_path in enumerate(images, 1):
        out_path = OUTPUT_DIR / (img_path.stem + ".png")

        if args.skip_done and out_path.exists():
            print(f"[{i}/{len(images)}] BỎ QUA (đã có): {img_path.name}")
            continue

        print(f"[{i}/{len(images)}] {img_path.name}")

        # Step 1: phân tích ảnh + detect brand
        try:
            description = analyze_image(img_path)
            brand = detect_brand(img_path, brands)
            if brand:
                fuzzy_query = f"{brand} {description}"
                print(f"  → Vision: {description} | Brand: {brand}")
            else:
                fuzzy_query = description
                print(f"  → Vision: {description}")
        except Exception as e:
            print(f"  ✗ Lỗi vision: {e}")
            unmatched_rows.append([img_path.name, "ERROR", str(e), 0, ""])
            continue

        # Step 2: match tên
        try:
            # Lấy ứng viên fuzzy để log (dùng fuzzy_query đã kết hợp brand)
            candidates = process.extract(fuzzy_query, product_names, scorer=fuzz.WRatio, processor=utils.default_process, limit=20)
            best_score = candidates[0][1] if candidates else 0
            best_candidate = candidates[0][0] if candidates else ""

            matched_name = match_product_name(fuzzy_query, product_names, brand=brand)
        except Exception as e:
            print(f"  ✗ Lỗi matching: {e}")
            unmatched_rows.append([img_path.name, description, str(e), 0, ""])
            continue

        if matched_name is None:
            print(f"  \033[33m⚠ KHÔNG KHỚP — fuzzy best: '{best_candidate}' ({best_score:.0f})\033[0m")
            unmatched_rows.append([img_path.name, description, brand or "", best_score, best_candidate])
            continue

        print(f"  ✓ Khớp: {matched_name}")

        # Step 3: tạo ảnh
        try:
            render_labeled_image(img_path, matched_name, out_path)
            print(f"  → Đã lưu: {out_path}")
        except Exception as e:
            print(f"  ✗ Lỗi render ảnh: {e}")
            unmatched_rows.append([img_path.name, description, brand or "", best_score, best_candidate])
            continue

        matched_rows.append([img_path.name, matched_name, description])

    # Ghi CSV
    if matched_rows:
        with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["filename", "matched_name", "description"])
            w.writerows(matched_rows)
        print(f"\nĐã ghi {len(matched_rows)} kết quả → {OUTPUT_CSV}")

    if unmatched_rows:
        with open(UNMATCHED_CSV, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["filename", "description", "brand_detected", "best_fuzzy_score", "best_fuzzy_candidate"])
            w.writerows(unmatched_rows)
        print(f"Không khớp {len(unmatched_rows)} ảnh → {UNMATCHED_CSV}")

    total = len(matched_rows) + len(unmatched_rows)
    print(f"\nHoàn thành: {len(matched_rows)}/{total} ảnh đã được nhận diện và tạo ảnh.")


if __name__ == "__main__":
    main()
